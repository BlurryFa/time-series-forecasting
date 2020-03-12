import pandas as pd
from pandas import  DataFrame, Series
import numpy as np
import random
import openpyxl


input_path = 'D:\\mnist_recog\\electric_analysis\\'
file_name = '2016-2018大户用电-筛选.xlsx'
# *************
# 用于填充的随机数
def data_reading(input_path=input_path):

    mouth = pd.PeriodIndex(start='2016-01', freq='M', periods=34)
    if not input_path is None:
        w = openpyxl.load_workbook(input_path + file_name)

        company_dict = []
        sheet = w.active

        for row in range(2, 49):
            company_name = sheet.cell(row=row, column=3).value
            company_list = []
            for column in range(7, 41):
                ele_consume = sheet.cell(row=row, column=column).value
                company_list.append(ele_consume)

            company_dict.append((company_name, company_list))

        time_series = DataFrame(dict(company_dict), index=mouth)

    else:
        data = []
        for i in range(731):
            data.append(random.uniform(1400,3000))
# **************

        index = pd.date_range('2017-01-01', '2019-01-01', freq='D')

        time_series = Series(data, index=index)
    time_series.index = time_series.index.to_timestamp()
    return  time_series

# if __name__ == '__main__':
#     data_reading()
